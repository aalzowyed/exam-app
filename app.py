import streamlit as st
import pandas as pd
import calendar
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.fills import GradientFill
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="منظم الاختبارات", layout="wide")

# تعيين اتجاه الصفحة من اليمين إلى اليسار
st.markdown("""
    <style>
        html, body, [data-testid="stAppViewContainer"] {
            direction: rtl;
            text-align: right;
        }
        .stButton button {
            width: 100%;
        }
    </style>
    """, unsafe_allow_html=True)

# إعدادات المواد والألوان
SUBJECTS_CONFIG = {
    "رياضيات": {"color": "#FF5733", "index": 0},
    "فيزياء": {"color": "#33FF57", "index": 1},
    "كيمياء": {"color": "#3357FF", "index": 2},
    "أحياء": {"color": "#F333FF", "index": 3},
    "لغة عربية": {"color": "#FFB833", "index": 4},
    "إنجليزي": {"color": "#33FFF5", "index": 5}
}

YEAR = 2026

# Initialize Session State
if "exams_data" not in st.session_state:
    st.session_state.exams_data = {}

if "page" not in st.session_state:
    st.session_state.page = "home"

st.title("📅 منظم الاختبارات الاحترافي - تصدير المخطط السنوي الملون")

# Navigation
col1, col2 = st.columns(2)
with col1:
    if st.button("إضافة اختبار جديد", key="btn_home"):
        st.session_state.page = "home"
        st.rerun()

with col2:
    if st.button("عرض المخطط السنوي 📅", key="btn_calendar"):
        st.session_state.page = "calendar"
        st.rerun()

st.divider()

# Page 1: Add Exam
if st.session_state.page == "home":
    st.subheader("إضافة اختبار جديد")
    
    col1, col2 = st.columns(2)
    with col1:
        selection_type = st.radio("نوع الإضافة:", ["يوم واحد", "فترة زمنية"], horizontal=True)
    
    with col2:
        subject = st.selectbox("المادة:", list(SUBJECTS_CONFIG.keys()))
    
    col1, col2, col3 = st.columns(3)
    with col1:
        date_start = st.date_input("تاريخ البداية:")
    
    if selection_type == "فترة زمنية":
        with col2:
            date_end = st.date_input("تاريخ النهاية:")
    else:
        date_end = date_start
    
    col1, col2, col3 = st.columns(3)
    with col1:
        hour = st.selectbox("الساعة:", [f"{i:02d}" for i in range(1, 13)], index=8)
    with col2:
        minute = st.selectbox("الدقيقة:", [f"{i:02d}" for i in range(0, 60, 5)])
    with col3:
        period = st.selectbox("AM/PM:", ["AM", "PM"])
    
    time_str = f"{hour}:{minute} {period}"
    
    if st.button("حفظ الاختبار", key="save_btn", use_container_width=True):
        if selection_type == "يوم واحد":
            dates_to_add = [date_start]
        else:
            dates_to_add = []
            if date_end < date_start:
                st.error("التاريخ غير صحيح!")
            else:
                curr = date_start
                while curr <= date_end:
                    dates_to_add.append(curr)
                    curr += timedelta(days=1)
        
        added_count = 0
        for d in dates_to_add:
            d_str = d.strftime('%d/%m/%Y')
            exam_info = {"type": subject, "time": time_str}
            
            if d_str in st.session_state.exams_data:
                if not any(ex['type'] == subject for ex in st.session_state.exams_data[d_str]):
                    st.session_state.exams_data[d_str].append(exam_info)
                    added_count += 1
            else:
                st.session_state.exams_data[d_str] = [exam_info]
                added_count += 1
        
        st.success(f"✅ تمت الإضافة لـ {added_count} يوم")
        st.rerun()

# Page 2: View Calendar and Export
elif st.session_state.page == "calendar":
    st.subheader("المخطط السنوي الملون")
    
    # Legend
    st.write("**مفتاح الألوان:**")
    legend_cols = st.columns(len(SUBJECTS_CONFIG))
    for idx, (sub, cfg) in enumerate(SUBJECTS_CONFIG.items()):
        with legend_cols[idx]:
            st.markdown(f'<div style="background-color: {cfg["color"]}; padding: 10px; border-radius: 5px; text-align: center; color: white; font-weight: bold;">{sub}</div>', unsafe_allow_html=True)
    
    st.divider()
    
    # Display Calendar
    months = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو", 
              "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]
    months_en = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    for m_idx in range(1, 13):
        with st.expander(f"📅 {months[m_idx-1]} ({months_en[m_idx-1]})"):
            first_day, num_days = calendar.monthrange(YEAR, m_idx)
            
            # Header with days
            cols = st.columns(7)
            days_names = ["الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت", "الأحد"]
            for col_idx, day_name in enumerate(days_names):
                with cols[col_idx]:
                    bg_color = "#FFE4C4" if col_idx >= 5 else "#f0f0f0"
                    st.markdown(f'<div style="background-color: {bg_color}; padding: 10px; text-align: center; font-weight: bold;">{day_name}</div>', unsafe_allow_html=True)
            
            # Calendar grid
            day_counter = 1
            for week in range(6):
                cols = st.columns(7)
                for col_idx in range(7):
                    with cols[col_idx]:
                        if week == 0 and col_idx < first_day:
                            st.write("")
                        elif day_counter <= num_days:
                            date_str = f"{day_counter:02d}/{m_idx:02d}/{YEAR}"
                            weekday = (first_day + day_counter - 1) % 7
                            
                            # Background color
                            bg = "#E8F8FF" if weekday == 4 else ("#FFF5E8" if weekday >= 5 else "white")
                            
                            # Check if there are exams
                            if date_str in st.session_state.exams_data:
                                exams = st.session_state.exams_data[date_str]
                                colors = [SUBJECTS_CONFIG[ex['type']]["color"] for ex in exams]
                                
                                html_content = f'<div style="background: linear-gradient(90deg, {", ".join(colors)}); padding: 15px; border-radius: 5px; text-align: center; color: white; font-weight: bold; cursor: pointer;" onclick="alert(\'{date_str}: {", ".join([ex["type"] + " " + ex["time"] for ex in exams])}\')"><strong>{day_counter}</strong></div>'
                            else:
                                html_content = f'<div style="background-color: {bg}; padding: 15px; border-radius: 5px; text-align: center; border: 1px solid #ddd;"><strong>{day_counter}</strong></div>'
                            
                            st.markdown(html_content, unsafe_allow_html=True)
                            
                            # Show details in collapsible
                            if date_str in st.session_state.exams_data:
                                with st.expander(f"تفاصيل {day_counter}/{m_idx}"):
                                    exams = st.session_state.exams_data[date_str]
                                    for exam in exams:
                                        st.write(f"🔹 **{exam['type']}** - الساعة: {exam['time']}")
                            
                            day_counter += 1
                        else:
                            st.write("")
    
    st.divider()
    
    # Export to Excel
    if st.button("تصدير مخطط Excel 📥", use_container_width=True, key="export_btn"):
        if not st.session_state.exams_data:
            st.warning("لا توجد بيانات للتصدير!")
        else:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "المخطط السنوي الملون"
                
                days_letters = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Headers
                for col in range(2, 45):
                    day_text = days_letters[(col-2) % 7]
                    cell = ws.cell(row=1, column=col, value=day_text)
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True)
                    if day_text == "Fri":
                        cell.fill = PatternFill(start_color="D1F2FF", fill_type="solid")
                    elif day_text in ["Sat", "Sun"]:
                        cell.fill = PatternFill(start_color="FFE4C4", fill_type="solid")
                    ws.column_dimensions[get_column_letter(col)].width = 4.5
                
                # Months
                months_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
                for m_idx, m_name in enumerate(months_names, start=1):
                    row_num = m_idx + 1
                    ws.cell(row=row_num, column=1, value=m_name).font = Font(bold=True)
                    ws.row_dimensions[row_num].height = 30
                    
                    first_day, num_days = calendar.monthrange(YEAR, m_idx)
                    for d in range(1, num_days + 1):
                        col_pos = d + first_day + 1
                        date_str = f"{d:02d}/{m_idx:02d}/{YEAR}"
                        weekday = (first_day + d - 1) % 7
                        
                        cell = ws.cell(row=row_num, column=col_pos, value=d)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = thin_border
                        
                        # Default color
                        default_color = "FFFFFF"
                        if weekday == 4:
                            default_color = "E8F8FF"
                        elif weekday >= 5:
                            default_color = "FFF5E8"
                        cell.fill = PatternFill(start_color=default_color, fill_type="solid")
                        
                        # Exam colors
                        if date_str in st.session_state.exams_data:
                            exams = st.session_state.exams_data[date_str]
                            num = len(exams)
                            if num == 1:
                                color = SUBJECTS_CONFIG[exams[0]['type']]["color"].replace("#", "")
                                cell.fill = PatternFill(start_color=color, fill_type="solid")
                            else:
                                colors = [SUBJECTS_CONFIG[ex['type']]["color"].replace("#", "") for ex in exams]
                                cell.fill = GradientFill(stop=colors)
                            cell.font = Font(bold=True)
                
                # Save to bytes
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.download_button(
                    label="📥 تحميل ملف Excel",
                    data=output,
                    file_name=f"exam_calendar_{YEAR}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                st.success("✅ تم إنشاء الملف بنجاح!")
            except Exception as e:
                st.error(f"❌ خطأ في التصدير: {e}")
    
    # Display saved exams
    if st.session_state.exams_data:
        st.divider()
        st.subheader("الاختبارات المحفوظة:")
        df_data = []
        for date_str, exams in sorted(st.session_state.exams_data.items()):
            for exam in exams:
                df_data.append({"التاريخ": date_str, "المادة": exam['type'], "الوقت": exam['time']})
        
        if df_data:
            df = pd.DataFrame(df_data)
            st.dataframe(df, use_container_width=True)
            
            if st.button("حذف جميع البيانات", key="clear_btn"):
                st.session_state.exams_data = {}
                st.success("تم حذف جميع البيانات")
                st.rerun()